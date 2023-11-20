from openpyxl import load_workbook
import csv


class Course:
    # When a new Course object is created, this method gets automatically called to set up these attributes with the
    # values provided during the object's instantiation. So for example if we create a new course (which would be a
    # course object then while creating it as;
    # new_course=Course("Thermodynamics","f111","12-3-2023"), now we don't have to specify its attributes separately each time
    def __init__(self, course_code, course_name, exam_date):
        self.course_code = course_code
        self.course_name = course_name
        self.exam_date = exam_date
        self.sections = []

    def get_all_sections(self):
        return self.sections

    def __str__(self):
        return f"Course Code: {self.course_code}\nCourse Name: {self.course_name}\nExam Date: {self.exam_date}"

    def populate_section(self):
        section_id = input("Enter section ID: ")
        day_slots = input("Enter day slots (e.g., 'Monday 9:00-11:00'): ")
        new_section = Section(section_id, day_slots)
        self.sections.append(new_section)
        print(f"Section {section_id} added successfully for {self.course_code}")




class Section:
    def __init__(self, section_id, day_slots, course):
        self.section_id = section_id
        self.day_slots = day_slots
        self.course = course

    def get_section_info(
            self):
        return f"Section ID: {self.section_id}, Day Slots: {self.day_slots}"


class Timetable:
    def __init__(self):
        self.courses = {}

    def enroll_subject(self, course):
        self.courses[course.course_code] = course
        print(f"{course.course_code} enrolled successfully.")

    def check_clashes(self):
        # Logic to check for clashes between sections or exams
        pass

    def export_to_csv(self, filename):
        with open(filename, 'w', newline='') as csvfile:
            fieldnames = ['Course Code', 'Course Name', 'Exam Date', 'Section ID', 'Day Slots']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            writer.writeheader()

            for course_code, course in self.courses.items():
                for section in course.get_all_sections():
                    writer.writerow({
                        'Course Code': course.course_code,
                        'Course Name': course.course_name,
                        'Exam Date': course.exam_date,
                        'Section ID': section.section_id,
                        'Day Slots': section.day_slots
                    })

        print(f"Timetable exported to {filename} successfully.")


def populate_courses_from_excel(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active

    courses = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        course_code, course_name, exam_date = row[:3]  # Assuming these columns exist in the Excel file
        new_course = Course(course_code, course_name, exam_date)
        courses.append(new_course)

        # Populate sections (For demonstration, I'm populating a default section for each course)
        new_course.populate_section()

    return courses


def main():
    timetable = Timetable()

    # # Manual enrollment of subjects to the timetable using CLI
    # course_code = input("Enter Course Code: ")
    # course_name = input("Enter Course Name: ")
    # exam_date = input("Enter Exam Date: ")
    #
    # new_course = Course(course_code, course_name, exam_date)
    # new_course.populate_section()
    # timetable.enroll_subject(new_course)

    # Populate courses from an Excel spreadsheet
    courses = populate_courses_from_excel("courses.xlsx")

    # Enroll subjects to the timetable
    for course in courses:
        timetable.enroll_subject(course)

    # Check for clashes
    timetable.check_clashes()

    # Export timetable to a CSV file
    timetable.export_to_csv("timetable.csv")


if __name__ == "__main__":
    # This block will only execute when this script is run directly
    # It won't run if this script is imported into another script
    # And that's why this __name__.... thing is used.
    main()
